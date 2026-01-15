"""
Export Service - Handles data export to CSV, Excel, and PDF formats.
Provides professional styled reports for management review.
"""

from sqlalchemy.orm import Session
from sqlalchemy import func
from typing import Tuple, Optional, List, Dict, Any
from datetime import date, datetime
import pandas as pd
import io
import logging

# Excel styling
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from app.models import Equipment, Intervention, SparePart, Technician, FailureMode, RPNAnalysis
from app.services.kpi_service import KPIService
from app.services.pdf_generator import PDFReportGenerator, ReportColors

logger = logging.getLogger(__name__)


# Excel styling constants
HEADER_FILL = PatternFill(start_color="3b82f6", end_color="3b82f6", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
BORDER_THIN = Border(
    left=Side(style='thin', color='E5E7EB'),
    right=Side(style='thin', color='E5E7EB'),
    top=Side(style='thin', color='E5E7EB'),
    bottom=Side(style='thin', color='E5E7EB')
)
ALT_ROW_FILL = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")


class ExportService:
    """Service class for data export operations"""
    
    @staticmethod
    async def export_interventions(
        db: Session,
        format: str,
        equipment_id: Optional[int] = None,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        type_panne: Optional[str] = None
    ) -> Tuple[bytes, str, str]:
        """
        Export interventions to CSV or Excel
        
        Args:
            db: Database session
            format: 'csv' or 'excel'
            equipment_id: Filter by equipment
            start_date: Filter start date
            end_date: Filter end date
            type_panne: Filter by failure type
        
        Returns:
            Tuple of (file_content, filename, media_type)
        """
        # Build query
        query = db.query(Intervention)
        
        if equipment_id:
            query = query.filter(Intervention.equipment_id == equipment_id)
        
        if start_date:
            query = query.filter(Intervention.date_intervention >= start_date)
        
        if end_date:
            query = query.filter(Intervention.date_intervention <= end_date)
        
        if type_panne:
            query = query.filter(Intervention.type_panne == type_panne)
        
        query = query.order_by(Intervention.date_intervention.desc())
        
        interventions = query.all()
        
        # Convert to DataFrame
        data = []
        for i in interventions:
            equipment = db.query(Equipment).filter(Equipment.id == i.equipment_id).first()
            
            data.append({
                'ID': i.id,
                'Equipment': equipment.designation if equipment else '',
                'Type de panne': i.type_panne or '',
                'Catégorie': i.categorie_panne or '',
                'Date intervention': i.date_intervention,
                'Date demande': i.date_demande,
                'Durée arrêt (h)': i.duree_arret,
                'Cause': i.cause or '',
                'Organe': i.organe or '',
                'Résumé': i.resume_intervention or '',
                'Coût matériel': i.cout_materiel,
                'Coût main d\'oeuvre': i.cout_main_oeuvre,
                'Coût total': i.cout_total,
                'Heures MO': i.nombre_heures_mo,
                'Statut': i.status.value
            })
        
        df = pd.DataFrame(data)
        
        # Export based on format
        if format == 'csv':
            # CSV with UTF-8 BOM for Excel compatibility
            output = io.BytesIO()
            df.to_csv(output, index=False, encoding='utf-8-sig')
            output.seek(0)
            
            filename = f"interventions_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
            media_type = "text/csv"
            
            return output.getvalue(), filename, media_type
        
        else:  # excel
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Interventions', index=False)
                
                # Auto-adjust column widths
                worksheet = writer.sheets['Interventions']
                for idx, col in enumerate(df.columns):
                    max_length = max(
                        df[col].astype(str).apply(len).max(),
                        len(col)
                    )
                    worksheet.column_dimensions[chr(65 + idx)].width = min(max_length + 2, 50)
            
            output.seek(0)
            
            filename = f"interventions_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            
            return output.getvalue(), filename, media_type
    
    @staticmethod
    async def export_equipment(
        db: Session,
        format: str,
        include_stats: bool = True
    ) -> Tuple[bytes, str, str]:
        """
        Export equipment list with optional statistics
        
        Args:
            db: Database session
            format: 'csv' or 'excel'
            include_stats: Include intervention statistics
        
        Returns:
            Tuple of (file_content, filename, media_type)
        """
        equipment_list = db.query(Equipment).order_by(Equipment.designation).all()
        
        data = []
        for eq in equipment_list:
            row = {
                'ID': eq.id,
                'Désignation': eq.designation,
                'Type': eq.type or '',
                'Localisation': eq.location or '',
                'Statut': eq.status.value,
                'Fabricant': eq.manufacturer or '',
                'Modèle': eq.model or '',
                'N° Série': eq.serial_number or '',
                'Date acquisition': eq.acquisition_date
            }
            
            if include_stats:
                interventions = db.query(Intervention).filter(
                    Intervention.equipment_id == eq.id
                ).all()
                
                total_interventions = len(interventions)
                total_downtime = sum(i.duree_arret for i in interventions)
                total_cost = sum(i.cout_total for i in interventions)
                
                row.update({
                    'Total interventions': total_interventions,
                    'Total arrêt (h)': round(total_downtime, 2),
                    'Coût total': round(total_cost, 2),
                    'MTTR (h)': round(total_downtime / total_interventions, 2) if total_interventions > 0 else 0
                })
            
            data.append(row)
        
        df = pd.DataFrame(data)
        
        if format == 'csv':
            output = io.BytesIO()
            df.to_csv(output, index=False, encoding='utf-8-sig')
            output.seek(0)
            
            filename = f"equipment_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
            media_type = "text/csv"
            
            return output.getvalue(), filename, media_type
        
        else:  # excel
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Equipment', index=False)
                
                worksheet = writer.sheets['Equipment']
                for idx, col in enumerate(df.columns):
                    max_length = max(df[col].astype(str).apply(len).max(), len(col))
                    worksheet.column_dimensions[chr(65 + idx)].width = min(max_length + 2, 50)
            
            output.seek(0)
            
            filename = f"equipment_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            
            return output.getvalue(), filename, media_type
    
    @staticmethod
    async def export_spare_parts(
        db: Session,
        format: str,
        low_stock_only: bool = False
    ) -> Tuple[bytes, str, str]:
        """
        Export spare parts inventory
        
        Args:
            db: Database session
            format: 'csv' or 'excel'
            low_stock_only: Export only low stock items
        
        Returns:
            Tuple of (file_content, filename, media_type)
        """
        query = db.query(SparePart)
        
        if low_stock_only:
            query = query.filter(SparePart.stock_actuel <= SparePart.seuil_alerte)
        
        query = query.order_by(SparePart.designation)
        
        parts = query.all()
        
        data = []
        for part in parts:
            data.append({
                'ID': part.id,
                'Désignation': part.designation,
                'Référence': part.reference,
                'Description': part.description or '',
                'Coût unitaire': part.cout_unitaire,
                'Stock actuel': part.stock_actuel,
                'Seuil alerte': part.seuil_alerte,
                'Unité': part.unite,
                'Fournisseur': part.fournisseur or '',
                'Délai livraison (j)': part.delai_livraison or '',
                'Stock bas': 'Oui' if part.stock_actuel <= part.seuil_alerte else 'Non'
            })
        
        df = pd.DataFrame(data)
        
        if format == 'csv':
            output = io.BytesIO()
            df.to_csv(output, index=False, encoding='utf-8-sig')
            output.seek(0)
            
            filename = f"spare_parts_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
            media_type = "text/csv"
            
            return output.getvalue(), filename, media_type
        
        else:  # excel
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Spare Parts', index=False)
                
                worksheet = writer.sheets['Spare Parts']
                for idx, col in enumerate(df.columns):
                    max_length = max(df[col].astype(str).apply(len).max(), len(col))
                    worksheet.column_dimensions[chr(65 + idx)].width = min(max_length + 2, 50)
            
            output.seek(0)
            
            filename = f"spare_parts_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            
            return output.getvalue(), filename, media_type
    
    @staticmethod
    async def export_kpi_report(
        db: Session,
        format: str,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        equipment_id: Optional[int] = None
    ) -> Tuple[bytes, str, str]:
        """
        Generate comprehensive KPI report
        
        Args:
            db: Database session
            format: 'excel' or 'pdf'
            start_date: Report start date
            end_date: Report end date
            equipment_id: Filter by equipment
        
        Returns:
            Tuple of (file_content, filename, media_type)
        """
        # Get all KPIs
        dashboard_kpis = KPIService.get_dashboard_kpis(db, start_date, end_date)
        
        if format == 'excel':
            output = io.BytesIO()
            
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                # Summary sheet
                summary_data = {
                    'Métrique': [
                        'MTBF (heures)',
                        'MTTR (heures)',
                        'Disponibilité (%)',
                        'Coût total',
                        'Coût matériel',
                        'Coût main d\'oeuvre',
                        'Total interventions',
                        'Interventions ouvertes',
                        'Equipements actifs',
                        'Techniciens actifs'
                    ],
                    'Valeur': [
                        dashboard_kpis['mtbf']['mtbf_hours'] if dashboard_kpis['mtbf'] else 0,
                        dashboard_kpis['mttr']['mttr_hours'] if dashboard_kpis['mttr'] else 0,
                        dashboard_kpis['availability']['availability_percentage'] if dashboard_kpis['availability'] else 0,
                        dashboard_kpis['cost_breakdown']['total_cost'] if dashboard_kpis['cost_breakdown'] else 0,
                        dashboard_kpis['cost_breakdown']['material_cost'] if dashboard_kpis['cost_breakdown'] else 0,
                        dashboard_kpis['cost_breakdown']['labor_cost'] if dashboard_kpis['cost_breakdown'] else 0,
                        dashboard_kpis['total_interventions'],
                        dashboard_kpis['open_interventions'],
                        dashboard_kpis['equipment_count'],
                        dashboard_kpis['technician_count']
                    ]
                }
                
                df_summary = pd.DataFrame(summary_data)
                df_summary.to_excel(writer, sheet_name='Résumé', index=False)
                
                # Failure distribution sheet
                if dashboard_kpis['failure_distribution']:
                    df_failures = pd.DataFrame(dashboard_kpis['failure_distribution'])
                    df_failures.to_excel(writer, sheet_name='Distribution pannes', index=False)
                
                # Format worksheets
                for sheet_name in writer.sheets:
                    worksheet = writer.sheets[sheet_name]
                    for idx, col in enumerate(worksheet.iter_cols()):
                        max_length = max(len(str(cell.value)) for cell in col)
                        worksheet.column_dimensions[chr(65 + idx)].width = min(max_length + 2, 50)
            
            output.seek(0)
            
            filename = f"kpi_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            
            return output.getvalue(), filename, media_type
        
        else:  # PDF format
            # Generate professional PDF report
            subtitle = None
            if start_date or end_date:
                start_str = start_date.strftime('%d/%m/%Y') if start_date else 'Début'
                end_str = end_date.strftime('%d/%m/%Y') if end_date else 'Aujourd\'hui'
                subtitle = f"Période: {start_str} - {end_str}"
            
            pdf = PDFReportGenerator("Rapport KPI - Maintenance", subtitle)
            pdf.add_header()
            
            # KPI Summary Section
            pdf.add_section("Indicateurs Clés de Performance")
            
            mtbf_val = dashboard_kpis['mtbf']['mtbf_hours'] if dashboard_kpis.get('mtbf') else 'N/A'
            mttr_val = dashboard_kpis['mttr']['mttr_hours'] if dashboard_kpis.get('mttr') else 'N/A'
            avail_val = f"{dashboard_kpis['availability']['availability_percentage']:.1f}%" if dashboard_kpis.get('availability') else 'N/A'
            
            # Overview info
            pdf.add_info_box(
                "Ce rapport détaille les performances de maintenance (MTBF, MTTR, Disponibilité) ainsi que l'analyse des coûts et des défaillances. "
                "Les indicateurs sont calculés sur la période sélectionnée.",
                "info"
            )

            pdf.add_kpi_row([
                {'value': f"{mtbf_val:.1f}h" if isinstance(mtbf_val, (int, float)) else mtbf_val, 'label': 'MTBF', 'description': 'Temps moyen entre pannes'},
                {'value': f"{mttr_val:.1f}h" if isinstance(mttr_val, (int, float)) else mttr_val, 'label': 'MTTR', 'description': 'Temps moyen réparation'},
                {'value': avail_val, 'label': 'Disponibilité', 'description': 'Taux opérationnel'},
            ])
            
            # Cost Summary
            pdf.add_section("Analyse des Coûts", "Répartition des coûts de maintenance par catégorie.")
            cost_data = dashboard_kpis.get('cost_breakdown', {})
            pdf.add_summary_box("Détail Financier", [
                ("Coût Total", f"{cost_data.get('total_cost', 0):,.2f} €"),
                ("Coût Matériel", f"{cost_data.get('material_cost', 0):,.2f} €"),
                ("Coût Main d'Oeuvre", f"{cost_data.get('labor_cost', 0):,.2f} €"),
            ])
            
            # Statistics
            pdf.add_section("Statistiques Générales", "Volumétrie des activités de maintenance.")
            pdf.add_summary_box("Métriques Opérationnelles", [
                ("Total Interventions", dashboard_kpis.get('total_interventions', 0)),
                ("Interventions Ouvertes", dashboard_kpis.get('open_interventions', 0)),
                ("Équipements Actifs", dashboard_kpis.get('equipment_count', 0)),
                ("Techniciens Actifs", dashboard_kpis.get('technician_count', 0)),
            ], two_columns=True)
            
            # Failure Distribution Table
            if dashboard_kpis.get('failure_distribution'):
                pdf.add_section("Distribution des Pannes", "Fréquence des types de pannes rencontrés.")
                headers = ['Type de Panne', 'Nombre d\'occurrences', 'Pourcentage']
                data = [
                    [item.get('type_panne', 'N/A'), item.get('count', 0), f"{item.get('percentage', 0):.1f}%"]
                    for item in dashboard_kpis['failure_distribution']
                ]
                pdf.add_table(headers, data, col_widths=[250, 150, 100])
            
            # Generate PDF
            pdf_content = pdf.generate()
            filename = f"kpi_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            media_type = "application/pdf"
            
            return pdf_content, filename, media_type
    
    @staticmethod
    async def export_amdec_report(
        db: Session,
        format: str,
        risk_level: Optional[str] = None,
        equipment_id: Optional[int] = None
    ) -> Tuple[bytes, str, str]:
        """
        Export AMDEC/RPN analysis report.
        
        Args:
            db: Database session
            format: 'excel' or 'pdf'
            risk_level: Filter by risk level (critical, high, medium, low)
            equipment_id: Filter by equipment
        
        Returns:
            Tuple of (file_content, filename, media_type)
        """
        # Subquery to get latest RPN ID for each failure mode (same as AMDECService)
        latest_rpn_id_subquery = db.query(
            func.max(RPNAnalysis.id).label('latest_id')
        ).group_by(
            RPNAnalysis.failure_mode_id
        ).subquery()

        # Build query for LATEST RPN analyses only
        query = db.query(RPNAnalysis).join(
            FailureMode, 
            RPNAnalysis.failure_mode_id == FailureMode.id
        ).join(
            Equipment,
            FailureMode.equipment_id == Equipment.id
        ).join(
            latest_rpn_id_subquery,
            RPNAnalysis.id == latest_rpn_id_subquery.c.latest_id
        ).filter(
            FailureMode.is_active == True
        )
        
        if equipment_id:
            query = query.filter(Equipment.id == equipment_id)
        
        if risk_level:
            # We'll calculate risk level in Python after fetching to stay consistent with stats
            pass

        query = query.order_by(RPNAnalysis.rpn_value.desc())
        rpn_results = query.all()
        
        # Filter by risk level
        def get_risk_level(rpn: int) -> str:
            if rpn >= 200: return "critical"
            elif rpn >= 100: return "high"
            elif rpn >= 50: return "medium"
            return "low"
            
        def get_risk_label(level: str) -> str:
            labels = {
                "critical": "Critique (>=200)",
                "high": "Élevé (100-199)",
                "medium": "Moyen (50-99)",
                "low": "Faible (<50)"
            }
            return labels.get(level, level)
        
        if risk_level:
            rpn_results = [r for r in rpn_results if get_risk_level(r.rpn_value) == risk_level]
        
        # Calculate summary statistics
        total_count = len(rpn_results)
        risk_counts = {"critical": 0, "high": 0, "medium": 0, "low": 0}
        for r in rpn_results:
            risk_counts[get_risk_level(r.rpn_value)] += 1
        
        if format == 'excel':
            output = io.BytesIO()
            
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                # Summary sheet
                summary_data = {
                    'Niveau de Risque': ['Critique (>=200)', 'Élevé (100-199)', 'Moyen (50-99)', 'Faible (<50)', 'Total'],
                    'Nombre': [risk_counts['critical'], risk_counts['high'], risk_counts['medium'], risk_counts['low'], total_count]
                }
                df_summary = pd.DataFrame(summary_data)
                df_summary.to_excel(writer, sheet_name='Résumé', index=False)
                
                # Detailed RPN analysis
                data = []
                for rpn in rpn_results:
                    fm = rpn.failure_mode
                    eq = fm.equipment if fm else None
                    data.append({
                        'Équipement': eq.designation if eq else 'N/A',
                        'Mode de Défaillance': fm.mode_name if fm else 'N/A',
                        'Cause': fm.failure_cause[:100] + '...' if fm and fm.failure_cause and len(fm.failure_cause) > 100 else (fm.failure_cause if fm else ''),
                        'Effet': fm.failure_effect[:100] + '...' if fm and fm.failure_effect and len(fm.failure_effect) > 100 else (fm.failure_effect if fm else ''),
                        'G (Gravité)': rpn.gravity,
                        'O (Occurrence)': rpn.occurrence,
                        'D (Détection)': rpn.detection,
                        'RPN': rpn.rpn_value,
                        'Niveau': get_risk_level(rpn.rpn_value).capitalize(),
                        'Action Corrective': rpn.corrective_action or '',
                        'Statut Action': rpn.action_status or 'pending'
                    })
                
                df_rpn = pd.DataFrame(data)
                df_rpn.to_excel(writer, sheet_name='Analyse RPN', index=False)
                
                # Apply styling
                for sheet_name in writer.sheets:
                    ws = writer.sheets[sheet_name]
                    # Style header row
                    for cell in ws[1]:
                        cell.fill = HEADER_FILL
                        cell.font = HEADER_FONT
                        cell.alignment = HEADER_ALIGN
                        cell.border = BORDER_THIN
                    # Style data and auto-width
                    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                        for cell in row:
                            cell.border = BORDER_THIN
                            if row_idx % 2 == 0:
                                cell.fill = ALT_ROW_FILL
                    # Auto column width
                    for column in ws.columns:
                        max_length = max(len(str(cell.value or '')) for cell in column)
                        ws.column_dimensions[column[0].column_letter].width = min(max_length + 2, 50)
            
            output.seek(0)
            filename = f"amdec_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            return output.getvalue(), filename, media_type
        
        else:  # PDF
            pdf = PDFReportGenerator(
                "Rapport AMDEC - Analyse des Risques", 
                f"Généré le {datetime.now().strftime('%d/%m/%Y')} - {total_count} analyses"
            )
            pdf.add_header()
            
            # Overview info box
            pdf.add_info_box(
                "Ce rapport présente l'analyse des modes de défaillance, de leurs effets et de leur criticité (AMDEC). "
                "Les actions correctives doivent être priorisées pour les risques critiques (RPN >= 200) et élevés (RPN >= 100).",
                "info"
            )
            
            # Risk Summary
            pdf.add_section("Synthèse par Niveau de Risque")
            pdf.add_kpi_row([
                {'value': str(risk_counts['critical']), 'label': 'Critique', 'description': 'Action immédiate requise'},
                {'value': str(risk_counts['high']), 'label': 'Élevé', 'description': 'Action planifiée requise'},
                {'value': str(risk_counts['medium']), 'label': 'Moyen', 'description': 'Surveillance recommandée'},
                {'value': str(risk_counts['low']), 'label': 'Faible', 'description': 'Acceptable'},
            ])
            
            # Top risks table
            if rpn_results:
                pdf.add_section("Détail des Risques Prioritaires", "Liste détaillée des modes de défaillance classés par priorité.")
                
                headers = ['Équipement / Composant', 'Mode de Défaillance', 'Causes & Effets', 'G / O / D', 'RPN', 'Action Recommandée']
                data = []
                # Use landscape mode for better width
                # Use all results up to 50 for detail, not just 20
                for rpn in rpn_results[:50]:
                    fm = rpn.failure_mode
                    eq = fm.equipment if fm else None
                    
                    # Consolidate equipment info
                    eq_info = f"{eq.designation if eq else 'N/A'}\n({eq.serial_number if eq else '-'})"
                    
                    # Consolidate causes and effects
                    cause_effect = f"Cause: {fm.failure_cause if fm else 'N/A'}\nEffet: {fm.failure_effect if fm else 'N/A'}"
                    
                    # Formatted scores
                    scores = f"G: {rpn.gravity}\nO: {rpn.occurrence}\nD: {rpn.detection}"
                    
                    data.append([
                        eq_info,
                        fm.mode_name if fm else 'N/A',
                        cause_effect,
                        scores,
                        str(rpn.rpn_value),
                        rpn.corrective_action or 'Aucune action définie'
                    ])
                
                # Custom column widths for landscape A4 (approx 800pts total width available)
                pdf.add_table(
                    headers, 
                    data, 
                    col_widths=[120, 100, 200, 50, 40, 180] 
                )
                
                if len(rpn_results) > 50:
                    pdf.add_text(f"... et {len(rpn_results) - 50} autres analyses non affichées dans ce rapport PDF.", "NoteText")
            
            pdf_content = pdf.generate(landscape_mode=True)
            filename = f"amdec_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            media_type = "application/pdf"
            return pdf_content, filename, media_type